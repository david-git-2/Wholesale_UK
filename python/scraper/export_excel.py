import json
import os
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

load_dotenv()

INPUT_JSON = Path("web/kbeauty/data/koba_data.json")

# Save this “business (URL only)” file here
OUTPUT_DIR = Path("outputs/business")  # change if you want

# Required sequence:
# SL, Image URL, Name, Stock Qty, Price, Commission %, Commission
COLUMNS = [
    ("sl", "SL"),
    ("image_url", "Image URL"),
    ("name", "Name"),
    ("stock_quantity", "Stock Qty"),
    ("price", "Price"),
    ("commission_percentage", "Commission %"),
    ("commission", "Commission"),
]

def autosize_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)  # cap width

def safe_filename_part(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return "products"
    out = []
    for ch in s:
        if ch.isalnum() or ch in ("-", "_"):
            out.append(ch)
        elif ch.isspace():
            out.append("_")
    cleaned = "".join(out).strip("_")
    return cleaned or "products"

def main():
    if not INPUT_JSON.exists():
        raise SystemExit(f"Input not found: {INPUT_JSON}")

    items = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    if not isinstance(items, list):
        raise SystemExit("web/kbeauty/data/koba_data.json must be a JSON array (list of products)")

    in_stock = [p for p in items if str(p.get("status", "")).lower() == "in_stock"]

    # Filename: <KOBA_PRODUCT_FILTER>_details_<YYYY-MM-DD>.xlsx
    raw_filter = os.getenv("KOBA_PRODUCT_FILTER", "products")
    first_part = raw_filter.strip().split()[0]
    product_filter = safe_filename_part(first_part)
    today = datetime.now().strftime("%Y-%m-%d")
    output_xlsx = OUTPUT_DIR / f"{product_filter}_details_{today}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "In Stock"

    # Header row
    ws.append([header for _, header in COLUMNS])

    # Data rows (with auto-increment SL)
    for idx, p in enumerate(in_stock, start=1):
        row = []
        for key, _ in COLUMNS:
            if key == "sl":
                row.append(idx)
            else:
                row.append(p.get(key, ""))
        ws.append(row)

    ws.freeze_panes = "A2"

    autosize_columns(ws)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    print(f"✅ Wrote {len(in_stock)} rows to {output_xlsx}")

if __name__ == "__main__":
    main()
