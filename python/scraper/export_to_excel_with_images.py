import io
import json
import os
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

load_dotenv()

INPUT_JSON = Path("docs/kbeauty/data/koba_data.json")
OUTPUT_DIR = Path("outputs/business_images")

# Read from ENV (with safe defaults)
COD_RATE = float(os.getenv("KOBA_COD_RATE", "0.01"))  # e.g. 0.01 = 1%
PACKING_WHITE_BOX = float(os.getenv("KOBA_PACKING_WHITE_BOX", "38"))
PACKING_WHITE_POLY = float(os.getenv("KOBA_PACKING_WHITE_POLY", "19"))
AWRC_FIXED = float(os.getenv("KOBA_AWRC_COMMISSION", "10"))  # fixed amount (NOT %)

# Columns (Commission % removed)
COLUMNS = [
    ("sl", "SL"),
    ("image", "Image"),
    ("name", "Name"),
    ("stock_quantity", "Stock Qty"),
    ("price", "Price"),
    ("commission", "Commission"),
    ("cod_charge", "COD Charge"),
    ("awrc_commission", "AWRC Commission"),
    ("packing_white_box", "Packing (White Box)"),
    ("final_commission_white_box", "Final Commission (White Box)"),
    ("packing_white_poly", "Packing (White Poly)"),
    ("final_commission_white_poly", "Final Commission (White Poly)"),
]

IMG_COL_INDEX = 2
IMG_WIDTH = 80
IMG_HEIGHT = 80


def autosize_columns(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def fetch_image_bytes(url: str, timeout=15) -> bytes | None:
    try:
        r = requests.get(url, timeout=timeout)
        r.raise_for_status()
        return r.content
    except Exception:
        return None


def safe_filename_part(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return "products"
    out = []
    for ch in s:
        if ch.isalnum() or ch in ("-", "_"):
            out.append(ch)
        elif ch.isspace():
            out.append("_")
    cleaned = "".join(out).strip("_")
    return cleaned or "products"


def main():
    if not INPUT_JSON.exists():
        raise SystemExit(f"Input not found: {INPUT_JSON}")

    items = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    if not isinstance(items, list):
        raise SystemExit("docs/kbeauty/data/koba_data.json must be a JSON array (list of products)")

    in_stock = [p for p in items if str(p.get("status", "")).lower() == "in_stock"]

    raw_filter = (os.getenv("KOBA_PRODUCT_FILTER") or "").strip()

# If empty, fallback to "products" for the filename
    first_part = raw_filter.split()[0] if raw_filter else "products"
    product_filter = safe_filename_part(first_part)
    today = datetime.now().strftime("%Y-%m-%d")
    output_xlsx = OUTPUT_DIR / f"{product_filter}_details_with_images_{today}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "In Stock (Images)"

    ws.append([header for _, header in COLUMNS])
    ws.freeze_panes = "A2"

    ws.column_dimensions[get_column_letter(IMG_COL_INDEX)].width = 16

    # key -> col index
    col_index = {key: i + 1 for i, (key, _) in enumerate(COLUMNS)}

    for idx, p in enumerate(in_stock, start=1):
        row_values = []
        for key, _ in COLUMNS:
            if key == "sl":
                row_values.append(idx)
            elif key == "image":
                row_values.append("")
            elif key in (
                "cod_charge",
                "awrc_commission",
                "packing_white_box",
                "final_commission_white_box",
                "packing_white_poly",
                "final_commission_white_poly",
            ):
                row_values.append("")
            else:
                row_values.append(p.get(key, ""))

        ws.append(row_values)
        r = ws.max_row

        # Column letters
        price_col = get_column_letter(col_index["price"])
        commission_col = get_column_letter(col_index["commission"])
        cod_col = get_column_letter(col_index["cod_charge"])
        awrc_col = get_column_letter(col_index["awrc_commission"])
        box_col = get_column_letter(col_index["packing_white_box"])
        final_box_col = get_column_letter(col_index["final_commission_white_box"])
        poly_col = get_column_letter(col_index["packing_white_poly"])
        final_poly_col = get_column_letter(col_index["final_commission_white_poly"])

        # Common charges
        ws[f"{cod_col}{r}"] = f"={price_col}{r}*{COD_RATE}"
        ws[f"{awrc_col}{r}"] = AWRC_FIXED

        # Packing values
        ws[f"{box_col}{r}"] = PACKING_WHITE_BOX
        ws[f"{poly_col}{r}"] = PACKING_WHITE_POLY

        # Two different finals
        ws[f"{final_box_col}{r}"] = f"={commission_col}{r}-({cod_col}{r}+{awrc_col}{r}+{box_col}{r})"
        ws[f"{final_poly_col}{r}"] = f"={commission_col}{r}-({cod_col}{r}+{awrc_col}{r}+{poly_col}{r})"

        # Optional formatting
        ws[f"{cod_col}{r}"].number_format = "0.00"
        ws[f"{final_box_col}{r}"].number_format = "0.00"
        ws[f"{final_poly_col}{r}"].number_format = "0.00"

        # Insert image
        url = (p.get("image_url") or "").strip()
        if url:
            img_bytes = fetch_image_bytes(url)
            if not img_bytes:
                ws.cell(row=r, column=IMG_COL_INDEX).value = url
            else:
                img_file = io.BytesIO(img_bytes)
                xl_img = XLImage(img_file)
                xl_img.width = IMG_WIDTH
                xl_img.height = IMG_HEIGHT
                cell_addr = f"{get_column_letter(IMG_COL_INDEX)}{r}"
                ws.add_image(xl_img, cell_addr)
                ws.row_dimensions[r].height = IMG_HEIGHT * 0.75

    autosize_columns(ws)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    print(f"✅ Wrote {len(in_stock)} rows with embedded images to {output_xlsx}")


if __name__ == "__main__":
    main()
