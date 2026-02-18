import os
import re
import json
import time
from collections import defaultdict
from datetime import datetime

import openpyxl

from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from googleapiclient.errors import HttpError
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow


ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DEFAULT_XLSX = os.path.join(ROOT_DIR, "pc", "data_file", "pc_data.xlsx")
DEFAULT_OUT_JSON = os.path.join(ROOT_DIR, "docs", "pc_data.json")
DEFAULT_OUT_IMAGES = os.path.join(ROOT_DIR, "pc", "out_images")

CREDS_DIR = os.path.join(ROOT_DIR, "pc", "credentials")
OAUTH_CLIENT_JSON = os.path.join(CREDS_DIR, "oauth_client.json")
TOKEN_JSON = os.path.join(CREDS_DIR, "token.json")

SCOPES = ["https://www.googleapis.com/auth/drive.file"]  # only files this app creates


def log(msg: str):
    print(msg, flush=True)


def safe_filename(s: str) -> str:
    s = str(s).strip()
    s = re.sub(r"[^\w\-\.]+", "_", s)
    return s[:120] if len(s) > 120 else s


def get_drive_service_oauth():
    os.makedirs(CREDS_DIR, exist_ok=True)

    creds = None
    if os.path.exists(TOKEN_JSON):
        log("🔐 Found token.json, using existing login...")
        creds = Credentials.from_authorized_user_file(TOKEN_JSON, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            log("🔄 Token expired, refreshing...")
            creds.refresh(Request())
        else:
            if not os.path.exists(OAUTH_CLIENT_JSON):
                raise FileNotFoundError(
                    f"Missing OAuth client file: {OAUTH_CLIENT_JSON}\n"
                    f"Download 'Desktop app' OAuth client JSON and save it there."
                )
            log("🌐 Opening browser for Google login/consent...")
            flow = InstalledAppFlow.from_client_secrets_file(OAUTH_CLIENT_JSON, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_JSON, "w") as token:
            token.write(creds.to_json())
        log(f"✅ Saved token: {TOKEN_JSON}")

    log("✅ Drive API client ready")
    return build("drive", "v3", credentials=creds)


def upload_to_drive(service, local_path: str, folder_id: str, make_public: bool = True) -> str:
    filename = os.path.basename(local_path)
    metadata = {"name": filename}
    if folder_id:
        metadata["parents"] = [folder_id]

    # Upload
    media = MediaFileUpload(local_path, resumable=True)
    created = service.files().create(body=metadata, media_body=media, fields="id").execute()
    file_id = created["id"]

    # Links
    direct_link = f"https://drive.google.com/uc?id={file_id}"          # best for <img src="">
    view_link = f"https://drive.google.com/file/d/{file_id}/view"      # fallback opens in browser

    if not make_public:
        return view_link

    # Make public (retry on Google internal errors)
    for attempt in range(1, 4):
        try:
            service.permissions().create(
                fileId=file_id,
                body={"type": "anyone", "role": "reader"},
            ).execute()
            return direct_link
        except HttpError as e:
            # Drive sometimes throws 500 internalError. Retry a couple times.
            log(f"⚠️  Public permission failed ({attempt}/3) for {filename}: {e}")
            time.sleep(1.5 * attempt)
        except Exception as e:
            log(f"⚠️  Public permission failed ({attempt}/3) for {filename}: {e}")
            time.sleep(1.5 * attempt)

    # If permissions still fail, continue without crashing
    log(f"⚠️  Could not make public after retries: {filename}. Using view link.")
    return view_link


def main():
    log("🚀 Starting PC export...\n")

    # ---- CONFIG ----
    EXCEL_PATH = DEFAULT_XLSX
    OUT_JSON_PATH = DEFAULT_OUT_JSON
    OUT_IMAGES_DIR = DEFAULT_OUT_IMAGES

    # Optional: paste a folder ID from *your* My Drive if you want uploads into a folder.
    # If left empty "", uploads go to My Drive root.
    DRIVE_FOLDER_ID = "1ADhsOk31vbtvv-o1alaLPLR0Vvjg0GrL"  # <-- optional

    HEADER_ROW = 4
    BARCODE_HEADER = "BARCODE"
    IMAGE_COLUMN_INDEX = 14
    SHEET_NAME = None

    MAKE_PUBLIC = True
    # ----------------

    log(f"📄 Excel: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel not found: {EXCEL_PATH}")

    os.makedirs(OUT_IMAGES_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(OUT_JSON_PATH), exist_ok=True)
    log(f"📁 Output images folder: {OUT_IMAGES_DIR}")
    log(f"🧾 Output JSON: {OUT_JSON_PATH}\n")

    log("📥 Loading workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sh = wb[SHEET_NAME] if SHEET_NAME else wb[wb.sheetnames[0]]
    log(f"✅ Using sheet: {sh.title}")

    max_col = sh.max_column
    max_row = sh.max_row
    log(f"📐 Sheet size: rows={max_row}, cols={max_col}")

    # Read headers
    log(f"🏷️ Reading headers from row {HEADER_ROW}...")
    headers = []
    for c in range(1, max_col + 1):
        v = sh.cell(HEADER_ROW, c).value
        headers.append(str(v).strip() if v is not None else f"col_{c}")

    # Find barcode col
    barcode_col = None
    for idx, h in enumerate(headers, start=1):
        if h.lower() == BARCODE_HEADER.lower():
            barcode_col = idx
            break
    if barcode_col is None:
        raise RuntimeError(f"Could not find header '{BARCODE_HEADER}' in row {HEADER_ROW}")

    log(f"✅ Found BARCODE column at index: {barcode_col}")

    # Read rows
    log("📦 Reading product rows...")
    start_data_row = HEADER_ROW + 1
    products_by_row = {}

    for r in range(start_data_row, max_row + 1):
        row_vals = [sh.cell(r, c).value for c in range(1, max_col + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        obj = {}
        for c, h in enumerate(headers, start=1):
            obj[h] = row_vals[c - 1]
        obj["_rowNumber"] = r
        products_by_row[r] = obj

    log(f"✅ Products loaded: {len(products_by_row)}")

    # Extract images
    log(f"🖼️ Finding embedded images anchored to column {IMAGE_COLUMN_INDEX}...")
    images = getattr(sh, "_images", [])
    log(f"🖼️ Total images detected in sheet: {len(images)}")

    img_by_row = {}
    for img in images:
        anchor = img.anchor._from  # 0-based
        row = anchor.row + 1
        col = anchor.col + 1

        if col != IMAGE_COLUMN_INDEX:
            continue
        if row not in products_by_row:
            continue

        img_bytes = img._data()
        ext = (getattr(img, "format", None) or "jpg").lower()
        if ext == "jpeg":
            ext = "jpg"

        barcode_val = products_by_row[row].get(headers[barcode_col - 1], "")
        barcode_str = safe_filename(barcode_val if barcode_val is not None else f"row_{row}")

        img_by_row[row] = (barcode_str, ext, img_bytes)

    log(f"✅ Images matched to product rows: {len(img_by_row)}")

    # Save images locally
    log("💾 Saving images locally...")
    used = defaultdict(int)
    local_path_by_row = {}

    for i, (row, (barcode_str, ext, img_bytes)) in enumerate(img_by_row.items(), start=1):
        used[barcode_str] += 1
        suffix = f"_{used[barcode_str]}" if used[barcode_str] > 1 else ""
        filename = f"{barcode_str}{suffix}.{ext}"
        local_path = os.path.join(OUT_IMAGES_DIR, filename)

        with open(local_path, "wb") as f:
            f.write(img_bytes)

        local_path_by_row[row] = local_path
        if i % 25 == 0 or i == len(img_by_row):
            log(f"   ...saved {i}/{len(img_by_row)}")

    log(f"✅ Local images saved: {len(local_path_by_row)}\n")

    # Upload to Drive
    log("☁️ Connecting to Google Drive...")
    service = get_drive_service_oauth()
    drive_url_by_row = {}

    log(f"⬆️ Uploading {len(local_path_by_row)} image(s) to Drive (MAKE_PUBLIC={MAKE_PUBLIC})...")
    for i, (row, path) in enumerate(local_path_by_row.items(), start=1):
        try:
            url = upload_to_drive(service, path, DRIVE_FOLDER_ID, make_public=MAKE_PUBLIC)
            drive_url_by_row[row] = url
        except Exception as e:
            log(f"❌ Upload failed for row={row}, file={os.path.basename(path)}: {e}")
            drive_url_by_row[row] = None

        if i % 10 == 0 or i == len(local_path_by_row):
            log(f"   ...uploaded {i}/{len(local_path_by_row)}")

    log(f"✅ Upload step done. URLs created: {sum(1 for v in drive_url_by_row.values() if v)}\n")

    # Build JSON
    log("🧾 Building JSON payload...")
    products = []
    for row, obj in products_by_row.items():
        out = dict(obj)
        out["imageUrl"] = drive_url_by_row.get(row)
        products.append(out)

    payload = {
        "meta": {
            "generatedAt": datetime.utcnow().isoformat() + "Z",
            "sourceFile": os.path.basename(EXCEL_PATH),
            "sheet": sh.title,
            "count": len(products),
            "imagesExtracted": len(local_path_by_row),
            "imagesUploaded": sum(1 for v in drive_url_by_row.values() if v),
            "makePublic": MAKE_PUBLIC,
        },
        "products": products
    }

    with open(OUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    log("\n✅ Done")
    log(f"- Products: {len(products)}")
    log(f"- Images extracted: {len(local_path_by_row)} -> {OUT_IMAGES_DIR}")
    log(f"- Images uploaded (with URL): {sum(1 for v in drive_url_by_row.values() if v)}")
    log(f"- JSON written: {OUT_JSON_PATH}")
    log(f"- Token saved: {TOKEN_JSON}")


if __name__ == "__main__":
    main()
